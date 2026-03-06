import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.db.models import QuerySet

class SpreadsheetGenerator:
    """
    Generates .xlsx byte streams from dynamically created Djagno QuerySets.
    """
    
    @staticmethod
    def _extract_headers(first_row: dict) -> list:
        return list(first_row.keys()) if first_row else ["No Data"]

    @classmethod
    def generate_excel(cls, queryset_iterable: list, report_title: str) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Report Data"
        
        # 1. Title Row
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"{report_title} - Generated {datetime.date.today().isoformat()}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        
        # 2. Header Row
        if not queryset_iterable:
            ws['A3'] = "No records found matching criteria."
            
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
            
        headers = cls._extract_headers(queryset_iterable[0])
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for col_idx, header_title in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            # Make titles readable ("created_at" -> "Created At")
            cell.value = str(header_title).replace('_', ' ').title()
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            
        # 3. Data Rows
        # (It's safe to load all into memory because Django's pagination should cap the massive sets, 
        # or we assume this background job has overhead clearance)
        for row_idx, row_dict in enumerate(queryset_iterable, 4):
            for col_idx, key in enumerate(headers, 1):
                raw_val = row_dict[key]
                
                # Coerce datetimes or UUIDs to strings for OpenPyXL safely
                if isinstance(raw_val, (datetime.datetime, datetime.date)):
                    val = raw_val.isoformat()
                elif hasattr(raw_val, 'hex'): # UUID check
                    val = str(raw_val)
                else:
                    val = raw_val
                    
                ws.cell(row=row_idx, column=col_idx, value=val)
                
        # Simple auto-sizing columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50)
            
        # 4. Save to byte-stream
        output = io.BytesIO()
        wb.save(output)
        
        return output.getvalue()
